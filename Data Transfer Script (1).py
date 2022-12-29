import os 
from openpyxl import Workbook, load_workbook 
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
import enum

#Note that for code to run, excel file must be closed before running code.
#Also note that directory must contain the excel file & this code open in the IDE for it to run.

class ExcelWriter:

#define a method for creating a workbook using the imported method, setting it active, and creating a worksheet
    
    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = file_name

#define a method for writing a header and subheader starting on the first two empty rows and merges the two rows together. 
#Method also sets fonts to specificed size, boldness, and name using imported module, as well as aligns center

    def write_header(self, header, subheader):
        self.worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
        self.worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header))
        self.worksheet['A1'] = header
        self.worksheet['A2'] = subheader
        self.worksheet['A1'].font = Font(name='Calibri', size=20, bold=True, italic=False)
        self.worksheet['A2'].font = Font(name='Calibri', size=12, bold=True, italic=False)

        self.worksheet["A1"].alignment = Alignment(horizontal='center')
        self.worksheet["A2"].alignment = Alignment(horizontal='center')

    def write_data(self, data):
        #for column orientation
        for i, row in enumerate(data):
            for j, cell in enumerate(row):
                self.worksheet.cell(row=i+3, column=j+1, value=cell)

        #OR for row orientation- need to fix this area of code and add a user input system to choose orientation of data arrangement
        #for i, row in enumerate(data):
            #for j, cell in enumerate(row):
                #self.worksheet.cell(row=j+1, column=i+3, value=cell)

    def save(self):
        self.workbook.save(self.file_name)
#This method simply saves the adjusted workbook

    def does_file_exist(file_name):
        return os.path.isfile(file_name)
#This method checks if a file exists with given name

# Example usage using a created array to transfer this automatically to an excel worksheet
data = [
    ['1st Trial', '2nd Trial', '3rd Trial', '4th Trial', '5th Trial'],
    [1, 15, 202, 240, 59],
    [2, 39, 25, 11, 101],
    [3, 25, 30, 21, 33],
    [25, 12, 33, 30, 44],
    [12, 245, 88, 29, 90]
]

#if os.path.exists(file_name):
    #Load the existing workbook
    #workbook = load_workbook(file_name)
    #Get the last row in the first sheet
    #worksheet = workbook[workbook.sheetnames[0]]
    #last_row = worksheet.max_row
    # Append the data to the worksheet

    
#calling the methods above with the given file/workbook name, and given header and subheader titles, then saves workbook
excel_writer = ExcelWriter("datafile.xlsx")
excel_writer.write_header('Force', 'N')
excel_writer.write_data(data)
excel_writer.save()